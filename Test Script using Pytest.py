import pytest
import openpyxl
from selenium import webdriver
from page_objects import LoginPage
from datetime import datetime

# Load the Excel data
workbook = openpyxl.load_workbook("login_test_data.xlsx")
sheet = workbook.active


@pytest.fixture(scope="module")
def driver():
    driver = webdriver.Chrome()  # or any other driver
    yield driver
    driver.quit()


def get_test_data():
    rows = list(sheet.iter_rows(values_only=True))
    return rows[1:]  # Exclude header row


@pytest.mark.parametrize("test_id,username,password,date,time_of_test,name_of_tester,test_result", get_test_data())
def test_login(driver, test_id, username, password, date, time_of_test, name_of_tester, test_result):
    driver.get("https://opensource-demo.orangehrmlive.com/web/index.php/auth/login")
    login_page = LoginPage(driver)

    login_page.enter_username(username)
    login_page.enter_password(password)
    login_page.click_login()

    try:
        WebDriverWait(driver, 10).until(
            EC.url_contains("/dashboard")
        )
        result = "Passed"
    except:
        result = "Failed"

    # Update the test result in Excel
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        if row[0].value == test_id:
            row[3].value = datetime.now().date()
            row[4].value = datetime.now().time()
            row[6].value = result
            break

    workbook.save("login_test_data.xlsx")
    assert result == "Passed", f"Test {test_id} failed"

